import streamlit as st
import easyocr
import pandas as pd
import re
import os
import io
import numpy as np
import pypdfium2 as pdfium
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Registry Tracker", layout="wide")

st.title("🎯 Precision Self-Exclusion Folder Extractor")
st.write("Production Version: Drag and drop files from your 'Selfban' folder to extract unseparated names and pristine identity numbers.")

# --- CACHE THE ENGINE TO PREVENT LAG ---
@st.cache_resource
def load_ocr_reader():
    """Loads the OCR engine once into memory so the app stays fast."""
    return easyocr.Reader(['en'], gpu=False)

try:
    reader = load_ocr_reader()
except Exception as e:
    st.error(f"OCR Engine Initialization Error: {e}")

def extract_perfect_data(file_bytes, file_name):
    """
    Directly extracts the unseparated full name and the 13-digit identity number
    from the exact GGB letter format layout structure provided.
    """
    base_name = os.path.splitext(file_name)[0].strip()
    full_name_clean = base_name.upper() if (base_name and not base_name.isspace()) else "UNKNOWN APPLICANT"
    id_number = "Not Found"
    
    try:
        pdf = pdfium.PdfDocument(file_bytes)
        
        # Target Page 1 where the official circular header resides
        page = pdf[0]
        bitmap = page.render(scale=2.5) 
        pil_img = bitmap.to_pil().convert('L')
        img_np = np.array(pil_img)
        
        ocr_results = reader.readtext(img_np, detail=0) 
        full_page_text = " ".join(ocr_results)
        
        # --- 1. IDENTITY NUMBER EXTRACTION ---
        id_regex_match = re.search(r'ID\s*No\.?\s*([\d\s\-]+)', full_page_text, re.IGNORECASE)
        if id_regex_match:
            raw_digits = re.sub(r'\D', '', id_regex_match.group(1))
            if len(raw_digits) >= 13:
                id_number = raw_digits[:13]
            elif raw_digits:
                id_number = raw_digits
                
        # Fallback: If OCR split the line layout, pull any continuous 13-digit block from the page
        if id_number == "Not Found":
            collapsed_text = full_page_text.replace(" ", "").replace("-", "").replace("–", "")
            global_id_match = re.search(r'\d{13}', collapsed_text)
            if global_id_match:
                id_number = global_id_match.group(0)

        # --- 2. FULL NAME EXTRACTION (UNSEPARATED) ---
        for line in ocr_results:
            line_upper = line.strip().upper()
            if "SELF" in line_upper and "BAN" in line_upper:
                standardized_line = line.replace("–", "split_here").replace("-", "split_here").replace(":", "split_here")
                parts = standardized_line.split("split_here")
                
                for part in parts:
                    part_upper = part.upper()
                    if any(title in part_upper for title in ["MR.", "MS.", "MRS.", "MR ", "MS "]):
                        name_segment = part
                        for title in ["Ms.", "Ms", "Mr.", "Mr", "Mrs.", "Mrs"]:
                            name_segment = name_segment.replace(title, "")
                        
                        if "ID" in name_segment.upper():
                            idx_id = name_segment.upper().find("ID")
                            name_segment = name_segment[:idx_id]
                            
                        name_segment = "".join([c for c in name_segment if not c.isdigit()])
                        name_final = name_segment.strip()
                        if name_final and len(name_final) > 2:
                            full_name_clean = name_final.upper()
                            break

    except Exception as e:
        pass

    return {
        "File Name": file_name,
        "Full Name": full_name_clean,
        "Identity Number": id_number
    }

def create_excel_download(dataframe):
    """Generates a beautifully formatted and padded Excel tracker file."""
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registry Database"
    ws.views.sheetView[0].showGridLines = True
    
    font_title = Font(name="Calibri", size=15, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11, bold=False, color="000000")
    
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    ws['A1'] = "Gauteng Gambling Board - Verified Registry Extraction"
    ws['A1'].font = font_title
    
    headers = ["File Name", "Full Name", "Identity Number"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    for row_idx, row_data in enumerate(dataframe.values, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = font_data
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
            if col_idx == 3:
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = '@'  # Enforces explicit text formatting to preserve leading zeros
            else:
                cell.alignment = Alignment(horizontal="left")
                
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[3].height = 24
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 18)
        
    wb.save(output)
    return output.getvalue()

# --- STREAMLIT UI SIDEBAR AND UPLOADER AREA ---
st.sidebar.header("Execution Settings")
st.sidebar.info("🌐 Web Folder Mode Active")

uploaded_files = st.file_uploader(
    "Drag and drop your compliance files or highlight everything in your 'Selfban' folder here:", 
    type=["pdf"], 
    accept_multiple_files=True
)

if uploaded_files:
    all_records = []
    if st.button("🚀 Run Targeted Extraction Pipeline"):
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for idx, uploaded_file in enumerate(uploaded_files):
            status_text.text(f"⏳ Processing profile [{idx+1}/{len(uploaded_files)}]: {uploaded_file.name}")
            
            file_bytes = uploaded_file.read()
            record = extract_perfect_data(file_bytes, uploaded_file.name)
            all_records.append(record)
            
            # Tick the progress bar smoothly
            progress_bar.progress((idx + 1) / len(uploaded_files))
            
        status_text.text("🎉 Extraction loop finished successfully!")
        
    if all_records:
        df = pd.DataFrame(all_records)
        st.success(f"Successfully finalized {len(all_records)} registry profiles!")
        
        st.subheader("📋 Processed Data Preview")
        st.dataframe(df, use_container_width=True)
        
        excel_data = create_excel_download(df)
        st.download_button(
            label="📥 Download Excel Spreadsheet",
            data=excel_data,
            file_name="Verified_Handwritten_Registry.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )